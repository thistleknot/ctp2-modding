"""Export MoM inventory and CSV surfaces into a single Excel workbook.

Purpose:
    Create a regenerable `.xlsx` that mirrors the scenario-owned option surfaces
    instead of relying on a hand-maintained spreadsheet.

Preconditions:
    - `openpyxl` must be importable in the active Python environment.
    - `dimension_inventory.md` must exist at the repository root.
    - The MoM control-plane CSVs must exist under `Scenarios\\mom\\tools\\momjr_csv`.

Failure modes:
    - Raises `FileNotFoundError` if the dimension inventory markdown is absent.
    - Raises `ImportError` if `openpyxl` is unavailable.
    - Raises `UnicodeDecodeError` only if a CSV cannot be decoded by the known
      fallback encodings.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TOOLS_DIR = Path(__file__).resolve().parent
ROOT_DIR = TOOLS_DIR.parents[2]
DIMENSION_INVENTORY = ROOT_DIR / "dimension_inventory.md"
DEFAULT_OUTPUT = ROOT_DIR / "Scenarios" / "mom" / "mom_dimension_inventory.xlsx"
CSV_ROOT = TOOLS_DIR / "momjr_csv"
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "latin-1")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Workbook path to write (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args()


def collect_csv_files(root: Path) -> list[Path]:
    """Return every MoM-owned control-plane CSV under momjr_csv."""
    return sorted(path for path in root.rglob("*.csv") if path.is_file())


def read_csv_rows(path: Path) -> list[list[str]]:
    """Read a CSV with the known encoding fallbacks used in this repo."""
    last_error: UnicodeDecodeError | None = None
    for encoding in CSV_ENCODINGS:
        try:
            with path.open(newline="", encoding=encoding) as handle:
                return [list(row) for row in csv.reader(handle)]
        except UnicodeDecodeError as exc:
            last_error = exc
    assert last_error is not None
    raise last_error


def parse_markdown_row(line: str) -> list[str]:
    """Split a markdown table row into trimmed cell values."""
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def is_separator_row(cells: list[str]) -> bool:
    """Return True when a markdown row is the --- alignment separator."""
    return bool(cells) and all(cell and set(cell) <= {"-", ":", " "} for cell in cells)


def parse_markdown_tables(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    """Extract markdown tables and their nearest heading labels."""
    if not path.exists():
        raise FileNotFoundError(path)

    lines = path.read_text(encoding="utf-8").splitlines()
    tables: list[tuple[str, list[str], list[list[str]]]] = []
    current_heading = "dimension_inventory"
    heading_counts: dict[str, int] = {}
    index = 0

    while index < len(lines):
        stripped = lines[index].strip()
        if stripped.startswith("#"):
            current_heading = stripped.lstrip("#").strip() or current_heading
            index += 1
            continue
        if not stripped.startswith("|"):
            index += 1
            continue

        block: list[str] = []
        start = index
        while index < len(lines) and lines[index].strip().startswith("|"):
            block.append(lines[index].strip())
            index += 1

        if len(block) < 2:
            continue
        header = parse_markdown_row(block[0])
        separator = parse_markdown_row(block[1])
        if not is_separator_row(separator):
            index = start + 1
            continue
        rows = [parse_markdown_row(line) for line in block[2:]]
        heading_counts[current_heading] = heading_counts.get(current_heading, 0) + 1
        title = current_heading
        if heading_counts[current_heading] > 1:
            title = f"{current_heading} {heading_counts[current_heading]}"
        tables.append((title, header, rows))

    return tables


def build_sheet_name(label: str, used_names: set[str]) -> str:
    """Create an Excel-safe, unique worksheet name."""
    base = "".join(ch if ch.isalnum() or ch in {"_", " "} else "_" for ch in label)
    base = "_".join(base.split()).strip("_") or "sheet"
    base = base[:31] or "sheet"
    name = base
    suffix = 2
    while name in used_names:
        stem = base[: max(1, 31 - len(str(suffix)) - 1)].rstrip("_")
        name = f"{stem}_{suffix}"
        suffix += 1
    used_names.add(name)
    return name


def style_header_row(worksheet, width_hint_rows: list[list[object]]) -> None:
    """Apply header styling, freeze panes, filter, and readable column widths."""
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for row in width_hint_rows:
            if column_index - 1 >= len(row):
                continue
            value = "" if row[column_index - 1] is None else str(row[column_index - 1])
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 48)


def write_tabular_sheet(worksheet, header: list[str], rows: list[list[object]]) -> None:
    """Write a simple table to a worksheet.

    Rows are written at their natural length: several control-plane CSVs are
    ragged (rows shorter AND longer than the header), and truncating to the
    header width silently loses the surplus cells on export.
    """
    worksheet.append(header)
    for row in rows:
        worksheet.append(list(row))
    style_header_row(worksheet, [header] + rows[:200])


def write_index_sheet(
    workbook: Workbook,
    sheet_specs: list[dict[str, object]],
    output_path: Path,
    csv_paths: list[Path],
    table_specs: list[tuple[str, list[str], list[list[str]]]],
    used_names: set[str],
) -> None:
    """Create the front-door sheet for the workbook."""
    sheet_name = build_sheet_name("index", used_names)
    worksheet = workbook.create_sheet(sheet_name, 0)
    worksheet["A1"] = "MoM Dimension Inventory Workbook"
    worksheet["A1"].font = TITLE_FONT
    worksheet["A3"] = "Output"
    worksheet["B3"] = str(output_path)
    worksheet["A4"] = "Dimension inventory source"
    worksheet["B4"] = str(DIMENSION_INVENTORY)
    worksheet["A5"] = "CSV root"
    worksheet["B5"] = str(csv_paths[0].parent if csv_paths else CSV_ROOT)
    worksheet["A6"] = "Workbook scope"
    worksheet["B6"] = "MoM-owned control-plane CSVs only (momjr_csv)"
    worksheet["A7"] = "CSV files included"
    worksheet["B7"] = len(csv_paths)
    worksheet["A8"] = "Markdown tables included"
    worksheet["B8"] = len(table_specs)
    worksheet["A10"] = "Workbook sheets"
    worksheet["A10"].font = SUBTITLE_FONT

    header = ["sheet", "kind", "source", "rows", "columns"]
    start_row = 11
    for offset, column_name in enumerate(header, start=1):
        cell = worksheet.cell(row=start_row, column=offset, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, spec in enumerate(sheet_specs, start=start_row + 1):
        worksheet.cell(row=row_index, column=1, value=spec["sheet"])
        worksheet.cell(row=row_index, column=2, value=spec["kind"])
        worksheet.cell(row=row_index, column=3, value=spec["source"])
        worksheet.cell(row=row_index, column=4, value=spec["rows"])
        worksheet.cell(row=row_index, column=5, value=spec["columns"])

    worksheet.freeze_panes = "A11"
    worksheet.auto_filter.ref = f"A10:E{start_row + len(sheet_specs)}"
    for column_letter, width in {"A": 28, "B": 16, "C": 72, "D": 10, "E": 10}.items():
        worksheet.column_dimensions[column_letter].width = width


def export_workbook(output_path: Path, csv_root: Path | None = None) -> tuple[Path, int]:
    """Build and save the workbook, returning the output path and sheet count.

    csv_root defaults to the MoM control plane (momjr_csv); pass another dir to
    export a different mod's csv set (dimension-inventory tables are skipped
    when the markdown is absent or when exporting a foreign csv root).
    """
    root = csv_root if csv_root is not None else CSV_ROOT
    csv_paths = collect_csv_files(root)
    include_inventory = csv_root is None and DIMENSION_INVENTORY.exists()
    table_specs = parse_markdown_tables(DIMENSION_INVENTORY) if include_inventory else []

    workbook = Workbook()
    workbook.remove(workbook.active)

    used_names: set[str] = set()
    sheet_specs: list[dict[str, object]] = []

    for title, header, rows in table_specs:
        sheet_name = build_sheet_name(f"table_{title}", used_names)
        worksheet = workbook.create_sheet(sheet_name)
        write_tabular_sheet(worksheet, header, rows)
        sheet_specs.append(
            {
                "sheet": sheet_name,
                "kind": "markdown-table",
                "source": f"{DIMENSION_INVENTORY.name} :: {title}",
                "rows": len(rows),
                "columns": len(header),
            }
        )

    for csv_path in csv_paths:
        rows = read_csv_rows(csv_path)
        if rows:
            header = rows[0]
            body = rows[1:]
        else:
            header = ["empty_csv"]
            body = []
        relative = csv_path.relative_to(root)
        label = "_".join(relative.with_suffix("").parts)
        sheet_name = build_sheet_name(label, used_names)
        worksheet = workbook.create_sheet(sheet_name)
        write_tabular_sheet(worksheet, header, body)
        sheet_specs.append(
            {
                "sheet": sheet_name,
                "kind": "csv",
                "source": str(relative),
                "rows": len(body),
                "columns": len(header),
            }
        )

    write_index_sheet(workbook, sheet_specs, output_path, csv_paths, table_specs, used_names)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path, len(workbook.sheetnames)


def main() -> None:
    args = parse_args()
    output_path, sheet_count = export_workbook(args.output)
    print(f"Wrote {output_path} ({sheet_count} sheet(s))")


if __name__ == "__main__":
    main()
