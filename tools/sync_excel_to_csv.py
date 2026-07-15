"""Sync control-plane CSVs from the mod inventory workbook (xlsx -> csv).

Purpose:
    Inverse of `export_mod_workbook.py`. Reads the workbook's index sheet to
    discover every csv-backed worksheet (kind == "csv") and writes each back to
    its source csv under the csv root. Markdown-table sheets are informational
    and never synced.

Preconditions:
    - `openpyxl` importable.
    - The workbook must contain the `index` sheet written by
      `export_mod_workbook.py` (falls back to matching sheet names against csv
      stems when the index is absent).

Guarantees:
    - Value-stable round-trip: export -> sync reproduces every csv cell value.
    - Per-file newline style and encoding are preserved for existing csvs
      (new csvs are written utf-8 with CRLF).
    - A worksheet whose header row disagrees with the on-disk csv header is
      REFUSED (schema drift) unless --force is given.

Failure modes:
    - Exits non-zero if the workbook is missing, a sheet named in the index is
      absent, or any sheet is refused for header drift (unless --force).
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
from pathlib import Path

import openpyxl

from export_mod_workbook import CSV_ROOT, DEFAULT_OUTPUT, CSV_ENCODINGS

INDEX_SHEET = "index"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook", type=Path, default=DEFAULT_OUTPUT,
        help=f"Workbook to read (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--csv-root", type=Path, default=CSV_ROOT,
        help=f"Destination csv root (default: {CSV_ROOT})",
    )
    parser.add_argument(
        "--check", action="store_true",
        help="Report which csvs would change (value-level diff); write nothing.",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Sync sheets even when their header row disagrees with the on-disk csv header.",
    )
    parser.add_argument(
        "--sheets", nargs="*", default=None,
        help="Restrict the sync to these sheet names (default: every csv-backed sheet).",
    )
    return parser.parse_args()


def sheet_to_csv_map(workbook, csv_root: Path) -> dict[str, Path]:
    """Map csv-backed worksheet names to their destination csv paths.

    Prefers the index sheet's provenance (kind == "csv", source == relpath);
    falls back to matching sheet names against csv stems under csv_root.
    """
    mapping: dict[str, Path] = {}
    if INDEX_SHEET in workbook.sheetnames:
        ws = workbook[INDEX_SHEET]
        header_row = None
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if header_row is None:
                if cells[:3] == ["sheet", "kind", "source"]:
                    header_row = cells
                continue
            if len(cells) >= 3 and cells[1] == "csv" and cells[0] and cells[2]:
                mapping[cells[0]] = csv_root / Path(cells[2])
        if mapping:
            return mapping

    # Fallback: name-match against existing csv files (no index sheet).
    stems = {"_".join(p.relative_to(csv_root).with_suffix("").parts): p
             for p in sorted(csv_root.rglob("*.csv"))}
    for name in workbook.sheetnames:
        if name in stems:
            mapping[name] = stems[name]
    return mapping


def _strip_trailing_empty(row: list[str]) -> list[str]:
    end = len(row)
    while end > 1 and not row[end - 1].strip():
        end -= 1
    return row[:end]


def sheet_rows(ws) -> list[list[str]]:
    """Worksheet -> [header, *body] as strings, canonical ragged form.

    openpyxl pads every row to the sheet's max width; the source CSVs are
    ragged, so the canonical form strips trailing empty cells per row (empty
    rows are dropped). Comparison and writing both use this form.
    """
    raw = [["" if c is None else str(c) for c in row]
           for row in ws.iter_rows(values_only=True)]
    raw = [r for r in raw if any(cell.strip() for cell in r)]
    return [_strip_trailing_empty(r) for r in raw]


def read_existing(path: Path) -> tuple[list[list[str]], str, str]:
    """Return (rows, encoding, newline) of an existing csv."""
    data = path.read_bytes()
    encoding = CSV_ENCODINGS[-1]
    for candidate in CSV_ENCODINGS:
        try:
            data.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    crlf = data.count(b"\r\n")
    lf = data.count(b"\n") - crlf
    newline = "\r\n" if crlf >= lf else "\n"
    text = data.decode(encoding)
    rows = [_strip_trailing_empty(["" if c is None else str(c) for c in r])
            for r in csv.reader(io.StringIO(text))]
    rows = [r for r in rows if any(cell.strip() for cell in r)]
    return rows, encoding, newline


def render_csv(rows: list[list[str]], newline: str) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator=newline)
    writer.writerows(rows)
    return buf.getvalue()


def main() -> int:
    args = parse_args()
    if not args.workbook.exists():
        print(f"ERROR: workbook not found: {args.workbook}")
        return 1

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    mapping = sheet_to_csv_map(wb, args.csv_root)
    if not mapping:
        print("ERROR: no csv-backed sheets found (missing index sheet and no name matches).")
        return 1

    wanted = args.sheets if args.sheets else sorted(mapping)
    refused: list[str] = []
    changed: list[str] = []
    synced = 0

    for name in wanted:
        if name not in mapping:
            print(f"REFUSED {name}: not a csv-backed sheet in this workbook")
            refused.append(name)
            continue
        if name not in wb.sheetnames:
            print(f"REFUSED {name}: sheet named in index but absent from workbook")
            refused.append(name)
            continue

        dest = mapping[name]
        rows = sheet_rows(wb[name])
        if not rows:
            print(f"SKIP    {name}: empty sheet")
            continue

        encoding, newline = "utf-8", "\r\n"
        existing_rows: list[list[str]] | None = None
        if dest.exists():
            existing_rows, encoding, newline = read_existing(dest)
            if existing_rows and existing_rows[0] != rows[0] and not args.force:
                print(f"REFUSED {name}: header drift vs {dest.name}")
                print(f"         csv : {existing_rows[0]}")
                print(f"         xlsx: {rows[0]}")
                refused.append(name)
                continue

        if existing_rows == rows:
            continue  # value-identical, leave bytes alone

        changed.append(name)
        if args.check:
            print(f"DRIFT   {name} -> {dest.relative_to(args.csv_root)}")
            continue

        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(render_csv(rows, newline).encode(encoding))
        synced += 1
        print(f"SYNCED  {name} -> {dest.relative_to(args.csv_root)} ({len(rows) - 1} rows)")

    if args.check:
        print(f"\n{len(changed)} sheet(s) differ from disk; {len(refused)} refused.")
    else:
        print(f"\n{synced} csv(s) written; {len(changed) - synced} skipped; {len(refused)} refused.")
    return 1 if refused else 0


if __name__ == "__main__":
    sys.exit(main())
